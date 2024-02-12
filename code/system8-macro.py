import argparse
from pathlib import Path
import pyautogui
import time
import os
import csv
import math
from enum import Enum
import sys
import warnings
from openpyxl import load_workbook
import tkinter as tk
from tkinter import filedialog

class Test(Enum):
    AMS_VI = "AMS-VI"
    AMS_Matrix = "AMS-Matrix"
    none = "None"


class Step:
    addStep     = (105, 1020)
    copyStep    = (185, 1020)
    delStep     = (267, 1020)
    firstStep   = ( 95, 202)
    lastStep    = ( 95, 652)

    incrStep    =   18
    maxStep     =   26

class Instrument:
    End         = (387, 62) # Where test window gets dragged to

class TF:
    new         = ( 14,  77)
    editMode    = ( 80, 115)
    setup_open  = (187, 1107)
    setup_close = (394,  92)
    instructions= (175, 830)


class ReportManager:
    AMS_VI      = ( 96, 232)
    AMS_Matrix  = ( 96, 248)

class AMS_VI:
    AMS_VI      = ( 12, 360) # Bring up test window
    probeSelect = (400, 110)
    pinSelect   = (400,135)

class AMS_Matrix:
    AMS_Matrix  = (12,390)
    pinSelect   = (360, 112)


class Preset_steps:
    def __init__(self):
        clickAt(TF.editMode)

    cur_step = 0
    
    def AMS_VI(self, text = "AMS VI", pins = 0, voltage = 5):
        
        createX(Test.AMS_VI)

        # Change probe / pins
        moveTo(AMS_VI.probeSelect)        
        if pins != 0:
            if "Probe" in pins:
                scroll(300)
                time.sleep(0.1)
                moveTo(AMS_VI.pinSelect)
                time.sleep(0.1)
                scroll(1000)
                time.sleep(0.1)
                if "Probes" in pins:
                    pins = pins.removesuffix(" Probes")
                    multi = math.floor(float(pins)) -1

                    scroll(-120*multi)
                    time.sleep(0.1)

                #print("PROBES")
            elif "Clip" in pins:
                pins = pins.removesuffix(" Clip")
                time.sleep(0.1)
                scroll(-300)
                time.sleep(0.1)
                #print("CLIP")
                moveTo(AMS_VI.pinSelect)
                time.sleep(0.1)
                scroll(5000)
            
                multi = math.floor(float(pins) / 2) -1

                scroll(-120*multi)
                time.sleep(0.1)

            else:
                print("Pins input not correct and has not changed")

        rename(text, self.cur_step)



    def AMS_Matrix(self, text = "AMS Matrix", pins = 0, voltage = 5):
       
        createX(Test.AMS_Matrix)
        
        #change pin number
        if pins != 0:
            if "Clip" in pins:
                pins = pins.removesuffix(" Clip")
            moveTo(AMS_Matrix.pinSelect)
            time.sleep(0.1)
            scroll(5000)
            
            multi = math.floor(float(pins) / 2) -1

            scroll(-120*multi)
            time.sleep(0.1)
            
        rename(text, self.cur_step)
    
    def IGNORE_STEP(self):
        return

    def Blank(self, text = "None"):
        
        createX(Test.none)
        rename(text, self.cur_step)

def moveTo(position):
    pyautogui.moveTo(position)

def clickAt(position):
    pyautogui.click(position)

def press(key):
    pyautogui.press(key)

def write(text):
    pyautogui.write(text)

def scroll(num):
    pyautogui.scroll(num)


probes = ["1 Probe", "2 Probes", "3 Probes", "4 Probes"]


def createX(test):

    match test:
        case Test.none:
            target = None
        case Test.AMS_VI:
            target = AMS_VI.AMS_VI
        case Test.AMS_Matrix:
            target = AMS_Matrix.AMS_Matrix
        case _:
            print("createX(" + str(test) + ") is not supported ")
            sys.exit()

    clickAt(Step.addStep)
    time.sleep(0.3)
    if(target):
        clickAt(target)
        time.sleep(0.5)

        moveWindow()

        time.sleep(0.1)
        
        reporting(test)

        time.sleep(0.1)


images = os.listdir("drag_images")

def moveWindow():

    location = None

    for image in images:
        location = pyautogui.locateOnScreen("drag_images/"+image, 0.4)
        if location != None:
            #print("found image")
            break

    if location is None:
        for image in images:
            location = pyautogui.locateOnScreen("drag_images/"+image, 1)
            if location != None:
                #print("found image")
                break

    # TODO: Add match case for different test windows
    start = (location[0] + 50, location[1] + 8)

    clickAt(start)
    time.sleep(0.5)
    pyautogui.dragTo(Instrument.End, duration=0.25)

rename_time = 0.5

def rename(text, stepNum = 1):

    if text is None:
        return

    if(stepNum >= Step.maxStep):
        #scroll down list
        target = Step.lastStep

    else:
        (x,y) = Step.firstStep
        y += Step.incrStep * (stepNum - 1)
        target = (x,y)
    

    clickAt(target)
    time.sleep(rename_time)
    clickAt(target)
    time.sleep(rename_time)
    clickAt(target)
    time.sleep(rename_time)
    clickAt(target)
    time.sleep(0.1)
    pyautogui.hotkey("ctrl", "a")
    write(text)
    press("enter")

def reporting(test):
    # click setup
    clickAt(TF.setup_open)


    # click appropriate test
    match test:
        case Test.AMS_VI:
            clickAt(ReportManager.AMS_VI)
        case Test.AMS_Matrix:
            clickAt(ReportManager.AMS_Matrix)
        case _:
            print("reporting(" + str(test) + ") is not supported")
            sys.exit()

    # close window
    clickAt(TF.setup_close)


def addInstructions(step, probePlus, probeMinus, notes):
    # click on instruction box

    clickAt(TF.instructions)
    if probePlus is not None:
        write("Probe + : ")
        write(probePlus)
        pyautogui.press("enter")
    if probeMinus is not None:
        write("Probe -	  :  ")
        write(probeMinus)
        pyautogui.press("enter")    
        pyautogui.press("enter")
    if notes is not None:
        write("Notes:")
        pyautogui.press("enter")
        write(notes)
    time.sleep(0.2)
    
    match step:
        case Test.AMS_VI:
            return

        case Test.AMS_Matrix:
            return

        case _:
            return

"""   
# Argument parsing
parser = argparse.ArgumentParser()
parser.add_argument("file")
args = parser.parse_args()
fileExists = Path(args.file).is_file()
if not fileExists:
    print("The target directory doesn't exist")
    raise SystemExit(1)

workbook = load_workbook(filename=args.file, read_only=True, data_only=True)
"""

"""
# file = "flows\BN3300-03 - System Monitor.xlsx"
# workbook = load_workbook(filename=file, read_only=True, data_only=True)
"""

root = tk.Tk()
root.withdraw()

# filepath = filedialog.askopenfilename(filetypes=["Excel files", "*.xlsx", "*.xls", "*.csv"])
filepath = filedialog.askopenfilename()
workbook = load_workbook(filename=filepath, read_only=True, data_only=True)


# PyAutoGUI settings 
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
pyautogui.FAILSAFE = True


ignoreSteps = ["RESERVED", "DO NOT EDIT"]
preset = Preset_steps()


#Load excel spreadsheet




worksheet = workbook.worksheets[0]
line_count = 0


notes = None
voltage = None
edit = None

for value in worksheet.iter_rows(values_only=True):
    line_count += 1

    # Extract column values from current row
    match len(value):
        case 6:
            num, name, step, pins, probePlus, probeMinus = value

        case 8:
            # print("8 wide input")
            num, name, step, pins, voltage, probePlus, probeMinus, notes = value
        case 9: 
            num, name, step, pins, voltage, probePlus, probeMinus, notes, edit = value
        case _:
            print("Input row length of " + str(len(value)) + " is not supported")
            sys.exit()

    # Remove unnecessary suffixes
    try:
        voltage = float(voltage.removesuffix("V pkpk"))
    except:
        voltage = 5

    print (line_count)

    if edit is not None:
            
        # Create the appropriate test
        if step in ignoreSteps:
            preset.IGNORE_STEP()

        elif step == "AMS-VI":
            preset.AMS_VI(name, pins, voltage)
            addInstructions(step, probePlus, probeMinus, notes)

        elif step == "AMS-Matrix":
            preset.AMS_Matrix(name, pins, voltage)
            addInstructions(step, probePlus, probeMinus, notes)
        # Add Test Instructions & Notes

        elif step == "None":
            preset.Blank(name)
            addInstructions(step, probePlus, probeMinus, notes)

#        preset.cur_step += 1

    else:
        preset.IGNORE_STEP()
    
    
    preset.cur_step += 1
    
print("\nTestflow complete!!. \nVerify all steps are correct and save.")

"""

row_list = []
for r in worksheet.rows:
    column = [cell.value for cell in r]
    row_list.append(column)

for row in row_list:
    print(row[0])



    line_count += 1
    if len(row) == 3:
        name, step, pins = row  
    
    elif len(row) < 3:
        print("Row " + line_count + " is too short")
        sys.exit()

    elif len(row) > 3:
        print("Row " + line_count + " is too long")
        sys.exit()
    #print(count, name, step)

    if step in ignoreSteps:
        preset.IGNORE_STEP()

    elif step == "AMS-VI":
        preset.AMS_VI(name, pins)

    elif step == "AMS-Matrix":
        preset.AMS_Matrix(name, pins)
"""

"""
steps = Preset_steps()

steps.AMS_VI()
print(steps.cur_step)
steps.AMS_Matrix()
print(steps.cur_step)

# Start - move all scrollables to the top 

"""